import datetime
import sys
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, PatternFill
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import Font



def load_playerslist(file_name, sheetname=None):
    wb = load_workbook(file_name)
    ws = wb[sheetname] if sheetname else wb.active
    players = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] and row[1]:
            players.append((row[0], row[1]))
    return players

def create_team_sheet(wb, sheetname, players, start_date, end_date):
    ws = wb.create_sheet(title=sheetname)

    center_align = Alignment(horizontal="center", vertical="center")
    rotated_align = Alignment(horizontal="left", vertical="bottom", text_rotation=90)

    saturday_fill = PatternFill(start_color="3399ff", end_color="3399ff", fill_type="solid")        # Saturday / Gameday bright blue
    excused_fill = PatternFill(start_color="ffff00", end_color="ffff00", fill_type="solid")         # "e" excused = kind of yellow
    excusedholiday_fill = PatternFill(start_color="ff8000", end_color="ff8000", fill_type="solid")  # "f" holiday = bright kind of yellow
    unexcused_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")       # "u" unexcused = red
    attendpractice_fill = PatternFill(start_color="6fdc6f", end_color="6fdc6f", fill_type="solid")  # "a" attend / present = green
    gameattend_fill = PatternFill(start_color="99ccff", end_color="99ccff", fill_type="solid")      # "s" Game attend = bright bright blue
    otherteam_fill = PatternFill(start_color="cc99ff", end_color="cc99ff", fill_type="solid")       # "t" Goes with other team = bright purpel

    practice_days = [0, 2]  # Monday and Wednesday
    gameday = 5             # Saturday
    header_row = 3          # where all the headers will be defined on row/line 3
    rotated_width_col = 4   # set the width to your requested width
    default_font = Font(name="Calibri", size=14, bold=False)

    # ------------------------------------
    # Headers
    # ------------------------------------
    header_font = Font(bold=True, size=14)
    ws["A1"] = sheetname
    ws["A1"].font = Font(bold=True, size=24)
    ws["A2"] = "Trainings bis heute:"
    ws["A2"].font = Font(bold=True, size=16)
    ws["A3"] = "Vorname"
    ws["B3"] = "Nachname"

    # ------------------------------------
    # Dates (Mon, Wed, Sat) from column C
    # ------------------------------------
    termine = []
    col = 3  # Column C
    first_letter_col = get_column_letter(col) # get first Column letter with dates
    current = start_date
    while current <= end_date:
        weekday = current.weekday()
    
        is_regular_practice = weekday in practice_days
        is_gameday = weekday == gameday
        if extra_day_start and extra_day_end:
            is_extra_day = (
                current.weekday() == 4 and extra_day_start <= current <= extra_day_end # 4 for Friday
            )
        else:
            is_extra_day = False

        if is_regular_practice or is_gameday or is_extra_day:
            zelle = ws.cell(row=header_row, column=col, value=current)
            zelle.number_format = "DD.MM.YYYY"
            zelle.alignment = rotated_align
            zelle.font = default_font
            termine.append((col, current))
            if current.weekday() == gameday:
                ws.cell(row=header_row, column=col).fill = saturday_fill
            col += 1
        current += datetime.timedelta(days=1)
    
    # get last Column letter with dates
    last_letter_col = get_column_letter(col-1)


    # practice dates
    practice_dates = [c for c, d in termine if d.weekday() in practice_days]
    game_dates = [c for c, d in termine if d.weekday() == gameday]

    # Count possible practice days until today (for B2)
    count_days_formula = (
        f'=SUMPRODUCT('
        f'--(WEEKDAY({first_letter_col}{header_row}:{last_letter_col}{header_row},2)=1) + '
        f'--(WEEKDAY({first_letter_col}{header_row}:{last_letter_col}{header_row},2)=3) + '
        f'--(WEEKDAY({first_letter_col}{header_row}:{last_letter_col}{header_row},2)=5), '
        f'--({first_letter_col}{header_row}:{last_letter_col}{header_row} <= TODAY())'
        f')'
    )
    ws["B2"] = count_days_formula

    # ------------------------------------
    # Additional columns for quota + games
    # ------------------------------------
    practicesquote_col = col
    game_col = col + 1
    ws.cell(row=header_row, column=practicesquote_col, value="Präsenz (%)").alignment=rotated_align
    ws.cell(row=header_row, column=game_col, value="Anz. Spiele").alignment=rotated_align

    # ------------------------------------
    # Participant data from row 4
    # ------------------------------------
    number_of_players=0
    for row_idx, (name, vorname) in enumerate(players, start=4):
        number_of_players += 1
        ws.cell(row=row_idx, column=1, value=name).font=default_font
        ws.cell(row=row_idx, column=2, value=vorname).font=default_font

        # Conditional formatting for each date cell
        for col_num, _ in termine:
            cell = ws.cell(row=row_idx, column=col_num).alignment=center_align
            col_letter = get_column_letter(col_num)
            cell_range = f"{col_letter}{row_idx}"
            cell.font = default_font

            ws.conditional_formatting.add(cell_range, CellIsRule(operator='equal', formula=['"u"'], fill=unexcused_fill))
            ws.conditional_formatting.add(cell_range, CellIsRule(operator='equal', formula=['"e"'], fill=excused_fill))
            ws.conditional_formatting.add(cell_range, CellIsRule(operator='equal', formula=['"a"'], fill=attendpractice_fill))
            ws.conditional_formatting.add(cell_range, CellIsRule(operator='equal', formula=['"s"'], fill=gameattend_fill))
            ws.conditional_formatting.add(cell_range, CellIsRule(operator='equal', formula=['"f"'], fill=excusedholiday_fill))
            ws.conditional_formatting.add(cell_range, CellIsRule(operator='equal', formula=['"t"'], fill=otherteam_fill))

            # set width for all iterated columns
            ws.column_dimensions[col_letter].width = rotated_width_col

        # Training area for this line
        if practice_dates:
            first_col_t = get_column_letter(practice_dates[0])
            last_col_t = get_column_letter(practice_dates[-1])
            practices_range = f"{first_col_t}{row_idx}:{last_col_t}{row_idx}"
            quote_formula = f'=IFERROR(COUNTIF({practices_range}, "a") / $B$2*100, "")'
            ws.cell(row=row_idx, column=practicesquote_col, value=quote_formula).font = default_font
        else:
            ws.cell(row=row_idx, column=practicesquote_col, value="").font = default_font

        # Game area for this line
        if game_dates:
            first_col_s = get_column_letter(game_dates[0])
            last_col_s = get_column_letter(game_dates[-1])
            game_range = f"{first_col_s}{row_idx}:{last_col_s}{row_idx}"
            game_formula = f'=COUNTIF({game_range}, "s")'
            ws.cell(row=row_idx, column=game_col, value=game_formula).font = default_font
        else:
            ws.cell(row=row_idx, column=game_col, value="").font = default_font
        

    # Total attendance
    total_start_row = len(players) +6
    ws.cell(row=total_start_row, column=1, value="TOTAL:").font = default_font
    ws.cell(row=total_start_row, column=2, value=number_of_players).font = default_font
    for total_col in range(column_index_from_string(first_letter_col),column_index_from_string(last_letter_col)+1):
        current_col=get_column_letter(total_col)
        num_attendance_formula = f'=COUNTIF({current_col}4:{current_col}{number_of_players+3},"a")' 
        ws.cell(row=total_start_row, column=total_col, value=num_attendance_formula).font = default_font

    # ------------------------------------
    # Insert legend below the list
    # ------------------------------------
    legend_start_row = len(players) + 8
    ws.cell(row=legend_start_row, column=1, value="Legende:").font = Font(name="Calibri", size=14, bold=True)
    ws.cell(row=legend_start_row + 1, column=1, value='a = anwesend',).fill=attendpractice_fill
    ws.cell(row=legend_start_row + 1, column=1).font = default_font
    ws.cell(row=legend_start_row + 2, column=1, value='e = entschuldigt abwesend').fill=excused_fill
    ws.cell(row=legend_start_row + 2, column=1).font = default_font
    ws.cell(row=legend_start_row + 3, column=1, value='f = ferien abwesend').fill=excusedholiday_fill
    ws.cell(row=legend_start_row + 3, column=1).font = default_font
    ws.cell(row=legend_start_row + 4, column=1, value='s = Spielteilnahme').fill=gameattend_fill
    ws.cell(row=legend_start_row + 4, column=1).font = default_font
    ws.cell(row=legend_start_row + 5, column=1, value='u = unentschuldigt abwesend').fill=unexcused_fill
    ws.cell(row=legend_start_row + 5, column=1).font = default_font
    ws.cell(row=legend_start_row + 6, column=1, value='t = Aufgebot anderes Team').fill=otherteam_fill
    ws.cell(row=legend_start_row + 6, column=1).font = default_font

    # -----------------------------------
    # Auto fit for columns and format
    # I used fix with width
    # -----------------------------------
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 20
    col_letter = get_column_letter(practicesquote_col)
    ws.column_dimensions[col_letter].width = rotated_width_col

    col_letter = get_column_letter(game_col)
    ws.column_dimensions[col_letter].width = rotated_width_col

    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=3, column=col)
        cell.font = header_font


### main 
# Input Verification - Usage message
def usage():
    print("Usage:")
    print("  python create-attendance-list.py <start_date> <end_date> [extra_day_start extra_day_end]")
    print("\nBeispiele:")
    print("  python create-attendance-list.py 2025-07-01 2025-09-30")
    print("  python create-attendance-list.py 2025-07-01 2025-09-30 2025-08-01 2025-08-31")
    sys.exit(1)

# Check: Did the user pass exactly two arguments?
if len(sys.argv) not in (3,5):
    print("❌ Error: Invalid number of arguments.")
    usage()

# Try to parse the dates
try:
    start_date = datetime.datetime.strptime(sys.argv[1], "%Y-%m-%d").date()
    end_date = datetime.datetime.strptime(sys.argv[2], "%Y-%m-%d").date()
except ValueError:
    print("❌ Error: Please enter the date values in the format YYYY-MM-DD.")
    usage()

# Logic check: Start date must be before or equal to end date
if start_date > end_date:
    print("❌ Error: Start date must not be after the end date.")
    usage()

# Optional: additional time range for additional practice day
extra_day_start = None
extra_day_end = None

if len(sys.argv) == 5:
    try:
        extra_day_start = datetime.datetime.strptime(sys.argv[3], "%Y-%m-%d").date()
        extra_day_end = datetime.datetime.strptime(sys.argv[4], "%Y-%m-%d").date()
    except ValueError:
        print("❌ Error: Please enter the date values in the format YYYY-MM-DD.")
        usage()
    
    if extra_day_start > extra_day_end:
        print("❌ Error: The additional range is invalid.")
        usage()

players_teamDa = load_playerslist("teamDa.xlsx")
players_teamDb = load_playerslist("teamDb.xlsx")

wb = Workbook()
# Remove the default “Sheet”
std = wb.active
wb.remove(std)

create_team_sheet(wb, "Team Da", players_teamDa, start_date, end_date)
create_team_sheet(wb, "Team Db", players_teamDb, start_date, end_date)

wb.save("attendance-list.xlsx")
