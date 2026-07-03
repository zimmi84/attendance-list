from openpyxl import load_workbook

wb = load_workbook("Jahresplan_Da-25_26.xlsx")

for ws in wb.worksheets:
    print(f"\n=== SHEET: {ws.title} ===")

    for row in ws.iter_rows():
        values = []
        for cell in row:
            values.append(f"{cell.coordinate}={cell.value}")
        print(" | ".join(values))
